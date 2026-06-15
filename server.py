#!/usr/bin/env python3
import json, os, glob, sys, webbrowser, threading
from http.server import HTTPServer, SimpleHTTPRequestHandler
import urllib.parse

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
    import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PORT = 8765

class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=BASE_DIR, **kwargs)

    def do_GET(self):
        if self.path == '/api/excel-files':
            # サブフォルダも含めて再帰スキャン、相対パスで返す
            names = []
            for root, dirs, files in os.walk(BASE_DIR):
                # 隠しフォルダをスキップ
                dirs[:] = [d for d in dirs if not d.startswith('.')]
                for f in sorted(files):
                    if f.lower().endswith(('.xlsx', '.xls')):
                        rel = os.path.relpath(os.path.join(root, f), BASE_DIR)
                        names.append(rel)
            self._send_json(sorted(names))

        elif self.path.startswith('/api/load?file='):
            raw = self.path.split('=', 1)[1]
            filename = urllib.parse.unquote(raw)
            # パストラバーサル対策
            filepath = os.path.realpath(os.path.join(BASE_DIR, filename))
            if not filepath.startswith(BASE_DIR) or not os.path.isfile(filepath):
                self.send_error(404, 'File not found')
                return
            words = []
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                ja = str(row[0] if row[0] is not None else '').strip()
                en = str(row[1] if row[1] is not None else '').strip()
                if ja and en:
                    words.append({'ja': ja, 'en': en})
            wb.close()
            self._send_json({'filename': filename, 'words': words})

        else:
            super().do_GET()

    def _send_json(self, data):
        body = json.dumps(data, ensure_ascii=False).encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', str(len(body)))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, *_):
        pass  # ログを静かにする

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()

    def do_POST(self):
        parsed = urllib.parse.urlparse(self.path)
        params = dict(urllib.parse.parse_qsl(parsed.query))
        length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(length) if length > 0 else b'{}'
        try:
            data = json.loads(body) if body else {}
        except Exception:
            data = {}

        if parsed.path == '/api/save':
            filename = urllib.parse.unquote(params.get('file', ''))
            if not filename:
                self.send_error(400, 'Missing file'); return
            filepath = os.path.realpath(os.path.join(BASE_DIR, filename))
            if not filepath.startswith(BASE_DIR):
                self.send_error(403, 'Forbidden'); return
            wb = openpyxl.Workbook()
            ws = wb.active
            for w in data.get('words', []):
                ws.append([w.get('ja', ''), w.get('en', '')])
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            wb.save(filepath)
            self._send_json({'ok': True})

        elif parsed.path == '/api/create':
            filename = data.get('filename', '').strip()
            if not filename:
                self.send_error(400, 'Missing filename'); return
            if not filename.lower().endswith('.xlsx'):
                filename += '.xlsx'
            subdir = os.path.join(BASE_DIR, '単語')
            filepath = os.path.realpath(os.path.join(subdir, os.path.basename(filename)))
            if not filepath.startswith(BASE_DIR):
                self.send_error(403, 'Forbidden'); return
            wb = openpyxl.Workbook()
            ws = wb.active
            for w in data.get('words', []):
                ws.append([w.get('ja', ''), w.get('en', '')])
            os.makedirs(subdir, exist_ok=True)
            wb.save(filepath)
            self._send_json({'ok': True, 'filename': os.path.relpath(filepath, BASE_DIR)})

        elif parsed.path == '/api/delete':
            filename = urllib.parse.unquote(params.get('file', ''))
            if not filename:
                self.send_error(400, 'Missing file'); return
            filepath = os.path.realpath(os.path.join(BASE_DIR, filename))
            if not filepath.startswith(BASE_DIR) or not os.path.isfile(filepath):
                self.send_error(404, 'Not found'); return
            os.remove(filepath)
            self._send_json({'ok': True})

        elif parsed.path == '/api/rename':
            filename = urllib.parse.unquote(params.get('file', ''))
            newname  = urllib.parse.unquote(params.get('newname', '')).strip()
            if not filename or not newname:
                self.send_error(400, 'Missing params'); return
            if not newname.lower().endswith(('.xlsx', '.xls')):
                newname += '.xlsx'
            filepath = os.path.realpath(os.path.join(BASE_DIR, filename))
            if not filepath.startswith(BASE_DIR) or not os.path.isfile(filepath):
                self.send_error(404, 'Not found'); return
            newpath = os.path.realpath(os.path.join(os.path.dirname(filepath), newname))
            if not newpath.startswith(BASE_DIR):
                self.send_error(403, 'Forbidden'); return
            os.rename(filepath, newpath)
            self._send_json({'ok': True, 'filename': os.path.relpath(newpath, BASE_DIR)})

        else:
            self.send_error(404, 'Not found')

def _open_browser():
    import time; time.sleep(0.6)
    webbrowser.open(f'http://localhost:{PORT}')

if __name__ == '__main__':
    # スマホ用にLAN IPも表示
    import socket
    try:
        lan_ip = socket.gethostbyname(socket.gethostname())
    except Exception:
        lan_ip = '(取得失敗)'

    print(f'✅ 単語帳サーバー起動中')
    print(f'   PC:    http://localhost:{PORT}')
    print(f'   スマホ: http://{lan_ip}:{PORT}  ← 同じWi-Fiなら使えます')
    print(f'終了: Ctrl+C')

    threading.Thread(target=_open_browser, daemon=True).start()
    try:
        HTTPServer(('', PORT), Handler).serve_forever()
    except KeyboardInterrupt:
        print('\n停止しました')
